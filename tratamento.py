#importações
import pandas as pd
from tqdm import tqdm
import openpyxl
import os
#import datetime as dt

def pega_tamanho_em_mb(caminho: str):
    return os.path.getsize(caminho) / (1024 * 1024)
    

def ler_excel(caminho):
    resultados = pd.DataFrame()
    tamanho_inicial = pega_tamanho_em_mb(caminho=caminho)
    print(f' O tamanho inicial é: {round(tamanho_inicial),2} MB')
    resultados['tamanho_inicial_mb'] = tamanho_inicial
    with tqdm(total=1, desc="Lendo Excel...") as pbar:
        try:
            df = pd.read_excel(caminho)
            pbar.update(1) # Indica que a leitura foi concluída
            return df
        except FileNotFoundError:
            print("Arquivo não encontrado.")
            return None
        except Exception as e:
            print(f"Ocorreu um erro ao ler o arquivo: {e}")
            return None
    resultados.to_csv('data_bronze/resultados1.csv')
    return df

#segunda tentativa de leitura
def ler_excel_com_progresso_openpyxl(caminho, chunk_size=1000):
    total_rows = 0
    try:
        workbook = openpyxl.load_workbook(caminho, read_only=True)
        sheet = workbook.active
        total_rows = sheet.max_row - 1  # Desconta o cabeçalho
        workbook.close()
        print(f"Total rows: {total_rows}")
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho}")
        return None
    except Exception as e:
        print(f"Erro ao obter o número de linhas de {caminho}: {e}")
        # Se falhar ao obter o número de linhas, lemos sem barra de progresso detalhada
        return pd.read_excel(caminho)

    all_data = []
    with tqdm(total=total_rows, unit='linha', desc="Lendo Excel") as pbar:
        workbook = openpyxl.load_workbook(caminho, read_only=True)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]  # Assume que a primeira linha é o cabeçalho
        print(f"Header: {header}")

        for i in range(2, total_rows + 2, chunk_size):
            print(f"Lendo chunk começando na linha: {i}")
            max_row = min(i + chunk_size - 1, total_rows + 1)
            chunk_data = []
            for row in sheet.iter_rows(min_row=i, max_row=max_row, values_only=True):
                chunk_data.append(row)

            if chunk_data:
                df_chunk = pd.DataFrame(chunk_data, columns=header)
                all_data.append(df_chunk)
                pbar.update(len(df_chunk))
                print(f"Chunk lido com {len(df_chunk)} linhas.")
            else:
                print(f"Nenhum dado lido no chunk começando na linha: {i}")

        workbook.close()

    if not all_data:
        print("Nenhum dado foi lido do arquivo.")

    df = pd.concat(all_data)
    return df

def processa_planilha(df):    
    #configura índice
    resultados = pd.DataFrame()
    qtde_bens = len(df.index)
    resultados['qtde_bens'] = qtde_bens
    resultados['qtde_inicial_colunas'] = len(df.columns)

    
    #proporção de valores nulos por coluna
    df_null = df.isnull().sum()/qtde_bens*100
    resultados['colunas_nulas_iniciais'] = df_null.sort_values(ascending=False).head(30)
    
    #checar se colunas de números de série existem na planilha
    cols_to_check = ['imei','n de serie', 'numero de serie',
                'numero de serie.1', 'numero de serie  ',
                 'num serie', 'placa', 'placa  ', 'placa vinculada',
                 'placa oficial', 'placa ','numero de serie.2']
    existing_serie_cols = [col for col in cols_to_check if col in df.columns]
    #resultados['colunas_existentes_numeros_serie'] = [existing_serie_cols]

    #checar se colunas de modelo existem na planilha
    cols_to_check = ['modelo', 'modelo  ', 'modelo    ', 'modelo1', 'modelo.1']
    existing_modelo_cols = [col for col in cols_to_check if col in df.columns]
    #resultados['existing_modelo_cols'] = [existing_modelo_cols]
    
    # checar se colunas de marca existem na planilha
    cols_to_check = ['marca','marca.1', 'marca1']
    existing_marca_cols = [col for col in cols_to_check if col in df.columns]
    #resultados['existing_marca_cols'] = [existing_marca_cols]
    
    #checar se colunas de tombo antigo existem na planilha
    cols_to_check = ['tombo antigo', 'tombo antigo.1']
    existing_tombo_antigo_cols = [col for col in cols_to_check if col in df.columns]
    #resultados['existing_tombo_antigo_cols'] = [existing_tombo_antigo_cols]

    # checar se colunas de especificações na planilha
    cols_to_check = ['observacao bloqueio', 'matriz', 'qtd de rodas',
       'acabamento da estrutura', 'altura', 'ano de fabricacao',
       'ano do modelo', 'aplicacao', 'bordas', 'calibre', 'calibre  ',
       'carga', 'data de validade', 'destino', 'genero', 'largura',
       'lote  numeros e letras sem espacos e caracteres especiais ',
       'material', 'material do assento e encosto',
       'material revestimento assento e encosto',
       'memoria de armazenamento', 'necessita ser substituido', 'nivel de protecao',
       'numero de chassis', 'numero de raias',
        'num serie  chassis',
       'ostensivo', 'profundidade', 'qtd de gavetas',
       'qtd de passageiros', 'qtd de portas', 'renavam',
       'sentido das raias', 'servidor responsavel', 'tamanho  novo ',
       'tipo de veiculo', 'alcance',
       'ano de fabricacao.1', 'aplicacao.1', 'blindagem', 'calibre.1',
       'capacidade', 'capacidade de tiros', 'combustivel',
       'compartimento cela', 'contraste', 'cor', 'cor predominante',
       'dimensao', 'espaco disco rigido', 'faixa de operacao',
       'frequencia', 'heavy duty', 'impedancia', 'interface',
       'largura de leitura', 'material.1', 'material da estrutura',
       'meio de aquisicao', 'numero de portas',
       'padrao de leitura', 'peso',
       'polegadas', 'potencia', 'potencia  cv ', 'qtd de canais',
       'qtd de nivel', 'qtd memoria ram', 'resolucao', 'revestimento',
       'tamanho da tela', 'taxa de transferencia', 'tensao',
       'tensao de alimentacao', 'tipo', 'tipo de identificacao',
       'tipo de propriedade', 'velocidade de varredura', 'voltagem',
       'zoom otico', 'nivel de protecao da placa', 'tipo do monitor',
        'carga.1',	'classe',	'portas',	'tanque',	'velocidade',
        'volume', 'bitola do pneu', 'numero do registro', 'qtde de canais',
        'nome da embarcacao', 'numero de registro','tipo de veiculo.1']
    existing_especificacoes_cols = [col for col in cols_to_check if col in df.columns]
    #resultados['existing_especificacoes_cols'] = [existing_especificacoes_cols]

    # criar coluna de serie que compilará os demais números de série
    df['serie_total'] = None
    df['modelo_total'] = None
    df['especificacoes'] = None
    df['tombo_antigo'] = None
    df['marca_total'] = None

    #lista_colunas_exibir = ['denominacao','serie_total', 'modelo_total', 'tombo_antigo', 'marca_total', 'especificacoes']

    #define as funções
    def create_especificacoes(row):
        especificacoes = {}
        for col in existing_especificacoes_cols:
            if col in row and not pd.isna(row[col]):
                especificacoes[col] = row[col]
        return especificacoes

    def compile_series(row, existing_serie_cols):
        lista_numero_series = []
        for col in existing_serie_cols:
            value = row[col]
            if not pd.isna(value) and value not in [" ","", ".", "..."]:
                lista_numero_series.append(str(value).strip())

        lista_numero_series = list(set(lista_numero_series))
        return ', '.join(lista_numero_series)

    def compile_modelo(row, existing_modelo_cols):
        lista_modelo = []
        for col in existing_modelo_cols:
            value = row[col]
            if not pd.isna(value) and value not in [" ","", ".", "..."]:
                lista_modelo.append(str(value).strip())

        lista_modelo = list(set(lista_modelo))
        return ', '.join(lista_modelo)

    def compile_marca(row, existing_marca_cols):
        lista_marca = []
        for col in existing_marca_cols:
            value = row[col]
            if not pd.isna(value) and value not in [" ","", ".", "..."]:
                lista_marca.append(str(value).strip())

        lista_marca = list(set(lista_marca))
        return ', '.join(lista_marca)

    def compile_tombo_antigo(row, existing_tombo_antigo_cols):
        lista_tombo_antigo = []
        for col in existing_tombo_antigo_cols:
            value = row[col]
            if not pd.isna(value) and value not in [" ","", ".", "..."]:
                for char in str(value):
                    value = str(value).lstrip('P')
                    value = str(value).lstrip('S')
                    value = str(value).lstrip('0')
                lista_tombo_antigo.append(str(value).strip())

        lista_tombo_antigo = list(set(lista_tombo_antigo))
        return ', '.join(lista_tombo_antigo)

    #chamar as funções
    df['especificacoes'] = df.apply(create_especificacoes, axis=1)
    df['serie_total'] = df.apply(compile_series, axis=1, args=(existing_serie_cols,))
    df['modelo_total'] = df.apply(compile_modelo, axis=1, args=(existing_modelo_cols,))
    df['marca_total'] = df.apply(compile_marca, axis=1, args=(existing_marca_cols,))
    df['tombo_antigo'] = df[existing_tombo_antigo_cols].apply(compile_tombo_antigo, axis=1, args=(existing_tombo_antigo_cols,))

    #exclui as colunas compiladas
    df.drop(columns=existing_tombo_antigo_cols, inplace=True)
    df.drop(columns=existing_modelo_cols, inplace=True)
    df.drop(columns=existing_serie_cols, inplace=True)
    df.drop(columns=existing_marca_cols, inplace=True)
    df.drop(columns=existing_especificacoes_cols, inplace=True)
    df.drop(columns='index', inplace=True)

    # dividir a célula e retornar a última parte após '-' para retitrar a sigla
    df['sigla'] = df['unidade responsavel material'].apply(lambda x: x.split('-')[-1])

    #trazer o tombo novo para a 1ª coluna (para o PROCV do excel)
    df = df.reindex(columns=['num tombamento'] + [col for col in df.columns if col != 'num tombamento'])
    qtde_colunas_depois = len(df.columns)
    resultados['qtde_final_colunas'] = qtde_colunas_depois
    resultados.to_csv('data_bronze/resultados2.csv')

    #Preencher campos vazios dos campos
    df['localidade'].fillna('Sem localidade', inplace=True)

    # preencher linhas vazias nas colunas
    df['ultimo levantamento'] = df['ultimo levantamento'].fillna("0000 / 2010")  # Repor NaN
    df['ano do levantamento'] = df['ultimo levantamento'].str.split('/').str[-1].str.strip().astype(int)     
    df['modelo_total'] = df['modelo_total'].fillna('Sem modelo')
    df['serie_total'] = df['serie_total'].replace('', 'Sem serial cadastrado')
    df['acautelado para'] = df['acautelado para'].replace('','Sem acautelamento')
    
    #transforma num tombamento em index
    df['index'] = df['num tombamento']

    #cria colunas para levantamento
    df['inventariado'] = df['ano do levantamento'].apply(lambda linha: 'sim' if linha == 2025 else None)

    #salvar csv de localidades únicas como lista
    localidades = df['localidade'].unique()
    localidades = pd.Series(localidades, name='localidade')
    pd.concat([localidades,pd.Series(['Nova localidade'])])
    localidades.to_csv('data_bronze/localidades.csv', index=False, header=False)
    
    return df

if __name__ == '__main__':
    CAMINHO = 'data/lista_bens.xlsx'
    df_lista_materiais = ler_excel(CAMINHO)
    #df_lista_materiais = ler_excel_com_progresso_openpyxl(CAMINHO, chunk_size=1000) #ValueError: No objects to concatenate
    
    df_processado = processa_planilha(df_lista_materiais)
    df_processado.to_csv('data_bronze/lista_bens-processado.csv')
    # gerar estatísticas dos stats após o processamento, comparando arquivo anterior e posterior (nº colunas, tamanho do arquivo, etc.)
    #guardar a quantidade de último levantamento e ano do levantamento e 
    # adicionar em arquivo de controle_levantamento (dia, unidade, quantidade de bens levantados) para poder gerar um gráfico de evolução

